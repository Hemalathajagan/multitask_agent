import json
import logging

logger = logging.getLogger(__name__)


async def edit_excel_file(file_path: str, sheet_name: str, updates_json: str) -> str:
    """Edit specific cells in an Excel file on the local filesystem.

    Args:
        file_path: Full path to the Excel file (e.g., C:/Users/user/Documents/report.xlsx).
        sheet_name: Name of the sheet to edit (use 'Sheet1' for default).
        updates_json: JSON string of cell updates, e.g. '{"A1": "Hello", "B2": 42, "C3": "=SUM(A1:A10)"}'.
    """
    try:
        from openpyxl import load_workbook
        from pathlib import Path

        path = Path(file_path)
        if not path.exists():
            return f"Error: File '{file_path}' not found."

        if not path.suffix.lower() in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
            return f"Error: File must be an Excel file (.xlsx). Got: {path.suffix}"

        updates = json.loads(updates_json)
        if not isinstance(updates, dict):
            return "Error: updates_json must be a JSON object with cell references as keys."

        wb = load_workbook(file_path)

        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            return f"Error: Sheet '{sheet_name}' not found. Available sheets: {available}"

        ws = wb[sheet_name]

        changes = []
        for cell_ref, value in updates.items():
            old_value = ws[cell_ref].value
            ws[cell_ref] = value
            changes.append(f"  {cell_ref}: '{old_value}' → '{value}'")

        try:
            wb.save(file_path)
        finally:
            wb.close()

        result = f"Successfully updated {len(changes)} cell(s) in '{file_path}' (sheet: {sheet_name}):\n"
        result += "\n".join(changes)
        logger.info(f"Excel edit: {file_path}, sheet={sheet_name}, {len(changes)} changes")
        return result

    except json.JSONDecodeError:
        return "Error: Invalid JSON in updates_json. Use format: {\"A1\": \"value\", \"B2\": 42}"
    except Exception as e:
        return f"Failed to edit Excel file: {str(e)}"
